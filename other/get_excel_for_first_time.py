import openpyxl
import json
import os

# Path to the Excel file

sprint = "137"
build = "137006"
sheets = ["memory trend","cpu trend"]

load = {'ControlPlane':'ControlPlane', 'SingleCustomer':'SingleCustomer','MultiCustomer':'MultiCustomer'}
load_type = load['SingleCustomer']

ROOT_PATH = "/Users/masabathulararao/Documents/Loadtest"

previous_excel_file_path = f'{ROOT_PATH}/generated_reports/{str(int(sprint)-1)}/{load_type}/previous_trend_excel.xlsx'
current_excel_file_path = f'{ROOT_PATH}/generated_reports/{sprint}/{load_type}/previous_trend_excel.xlsx'

os.makedirs(os.path.dirname(previous_excel_file_path), exist_ok=True)
if not os.path.exists(previous_excel_file_path):
    prev_workbook = openpyxl.Workbook()
    default_sheet = prev_workbook.active
    prev_workbook.remove(default_sheet)
    sheet1 = prev_workbook.create_sheet("memory trend")
    sheet2 = prev_workbook.create_sheet("cpu trend")
    prev_workbook.save(previous_excel_file_path)


# Input dictionary with keys and values

current_build_data_path = f'{ROOT_PATH}/generated_reports/{sprint}/{load_type}/mem_cpu_comparision_data.json'
with open(current_build_data_path, 'r') as file:
        current_data = json.load(file)

current_memory,current_cpu = current_data['memory'] , current_data['cpu']

def create_dict(current,unit,tag):
    data_dict=dict()
    for container in current.keys():
        for host in current[container]:
            if container=='Host':
                name = f'{tag} Used by {host}'
                data_dict[name] = f'{current[container][host]["percentage"]:.2f}% '
            else:
                name=f'{tag} used by {container} {host}'
                data_dict[name] = f'{current[container][host]["percentage"]:.2f}% '
            if current[container][host][unit]:
                 data_dict[name] += f"({current[container][host][unit]:.2f} {unit})"
                 
    return data_dict

memory_data_dict=create_dict(current_memory,"GB","memory")
cpu_data_dict=create_dict(current_cpu,"cores","cpu")

workbook = openpyxl.load_workbook(previous_excel_file_path)

def add_column(sheet_name,data_dict):
    sheet = workbook[sheet_name]
    col = sheet.max_column +1 
    sheet.cell(row=1, column=col, value=build)
    
    for key, value in data_dict.items():
        found=False
        row = sheet.max_row +1
        for row_num in range(2, row):
            cell_value = sheet.cell(row=row_num, column=1).value
            if cell_value is not None and sheet.cell(row=row_num, column=1).value.lower() == key.lower():
                found=True
                print(value)
                sheet.cell(row=row_num, column=col, value=value.lower())
                break  # Exit the loop after finding a match

        if not found:
            sheet.cell(row=row, column=1, value=key.lower())
            sheet.cell(row=row, column=col, value=value)

# Save the modified Excel file

add_column("memory trend",memory_data_dict)
add_column("cpu trend",cpu_data_dict)

workbook.save(current_excel_file_path)
