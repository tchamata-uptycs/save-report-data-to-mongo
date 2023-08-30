import openpyxl
import json
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

class excel_update:
    def __init__(self,sprint,build,load_type,previous_excel_file_path,current_excel_file_path,current_build_data_path,fetch_prev_build_data_path):
        self.previous_excel_file_path=previous_excel_file_path
        self.current_excel_file_path=current_excel_file_path
        self.current_build_data_path=current_build_data_path
        # self.overall_save_current_build_data_path=overall_save_current_build_data_path
        self.prev_build_data_path=fetch_prev_build_data_path
        # self.overall_prev_build_data_path=overall_fetch_prev_build_data_path
        self.sprint=sprint
        self.build=build
        self.load_type = load_type
    def create_dict_cpu_mem(self,current,prev,unit,tag):
        data_dict=dict()
        for container in current.keys():
            for host in current[container]:
                if container=='Host':
                    name = f'{tag} Used by {host}'
                else:
                    name=f'{tag} used by {container} {host}'
                data_dict[name] = [f'{round(current[container][host]["percentage"],2)}%']

                if current[container][host][unit]:
                    # data_dict[name].append(round(current[container][host][unit],2))
                    data_dict[name][0] += f'({round(current[container][host][unit],2)} {unit})'

                # try:data_dict[name].append(round(-prev[container][host][unit]+current[container][host][unit] , 2))
                # except:pass
                # try:data_dict[name].append(round((-prev[container][host][unit]+current[container][host][unit])*100/prev[container][host][unit] , 2))
                # except:pass
        return data_dict
    def create_dict_overall(self,current,prev,unit,tag):
        data_dict=dict()
        for node_type in current.keys():
            name=f'Average {tag} used by {node_type}'
            data_dict[name] = [round(current[node_type][unit],2)]
            try:data_dict[name].append(round(-prev[node_type][unit]+current[node_type][unit],2))
            except Exception as e:
                print("Error:", type(e).__name__, "-", str(e))
            try:data_dict[name].append(round((-prev[node_type][unit]+current[node_type][unit])*100/prev[node_type][unit],2))
            except Exception as e:
                print("Error:", type(e).__name__, "-", str(e))
        return data_dict
    def create_dict_container_wise(self,current,prev,unit,tag):
        data_dict=dict()
        for container in current.keys():
            for host in current[container]:
                name=f'{tag} used by {container} {host}'
                data_dict[name] = [round(current[container][host][unit],2)]
                try:data_dict[name].append(round(-prev[container][host][unit]+current[container][host][unit] , 2))
                except Exception as e:
                    # data_dict[name].append(round(current[container][host][unit] , 2))
                    print("Error:", type(e).__name__, "-", str(e))
                try:data_dict[name].append(round((-prev[container][host][unit]+current[container][host][unit])*100/prev[container][host][unit] , 2))
                except Exception as e:
                    # data_dict[name].append(0)
                    print("Error:", type(e).__name__, "-", str(e))
        return data_dict
    
    def create_dict_node_apps(self,current,prev,unit,tag):
        data_dict=dict()
        for container in current:
            name=f'{tag} used by {container}'
            data_dict[name] = [round(current[container]["percentage"],2)]
        return data_dict

    def create_complete_dict(self,data,prev):
        data_dict=dict()
        for tag in data:
            if tag == 'Memory':
                name = f"Complete {tag} usage in  GB"
            elif tag=='CPU':
                name = f"Complete {tag} usage in  cores"
            else:
                name = f"Complete {tag} usage"
            data_dict[name]  = [round(data[tag] , 2)]
            try:data_dict[name].append(round(data[tag]-prev[tag] , 2))
            except Exception as e:
                    print("Error:", type(e).__name__, "-", str(e))
            try:data_dict[name].append(round((data[tag]-prev[tag])*100/prev[tag] , 2))
            except Exception as e:
                    print("Error:", type(e).__name__, "-", str(e))
        return data_dict
    
    def update_col_width(self,sheet,value_width,col):
        col_letter = get_column_letter(col)
        if col_letter not in sheet.column_dimensions:
            sheet.column_dimensions[col_letter].width = value_width
        else:
            current_width = sheet.column_dimensions[col_letter].width
            if value_width > current_width:
                sheet.column_dimensions[col_letter].width = value_width

    def create_kafka_dict(self,lst):
        data_dict=dict()
        for topic in lst:
            data_dict[topic] = [topic]
        return data_dict
    
    def add_row(self,sheet,value,row_num,temp_col,found):
        if len(value)>2:
            for val in value[:-2]:
                try:cell_val=abs(val)
                except Exception as e:
                    print("Error:", type(e).__name__, "-", str(e))
                    cell_val=val
                new_cell = sheet.cell(row=row_num, column=temp_col, value=cell_val)
                if not found:
                    new_cell.fill = PatternFill(start_color="b8f1ff", end_color="b8f1ff", fill_type="solid")  # blue fill
                temp_col+=1
            
            for val in value[-2:]:
                try:cell_val=abs(val)
                except Exception as e:
                    print("Error:", type(e).__name__, "-", str(e))
                    cell_val=val
                cell = sheet.cell(row=row_num, column=temp_col, value=cell_val)
                if val > 0:
                    cell.fill = PatternFill(start_color="F0C0C0", end_color="F0C0C0", fill_type="solid")  # Red fill
                elif val <0:
                    cell.fill = PatternFill(start_color="C0F0CF", end_color="C0F0CF", fill_type="solid")  # Green fill
                temp_col+=1
        else:
            for val in value:
                try:cell_val=abs(val)
                except Exception as e:
                    print("Error:", type(e).__name__, "-", str(e))
                    cell_val=val
                new_cell = sheet.cell(row=row_num, column=temp_col, value=cell_val)
                if not found:
                    new_cell.fill = PatternFill(start_color="b8f1ff", end_color="b8f1ff", fill_type="solid")  # blue fill
                temp_col+=1
        return sheet
    

    def add_columns(self,sheet_name,data_dict,workbook):
        sheet = workbook[sheet_name]
        col = sheet.max_column +1 
        while(col>2 and sheet.cell(row=1,column=col-1).value is None):
            col-=1
            sheet.delete_cols(col)

        build_cell = sheet.cell(row=1, column=col, value=int(self.build))
        build_cell.font = Font(bold=True)
        metric_cell = sheet.cell(row=1, column=1, value="Metric")
        metric_cell.font = Font(bold=True)
        sheet.freeze_panes = sheet.cell(row=2, column=2)
        for key, value in data_dict.items():
            found=False
            row = sheet.max_row +1
            for row_num in range(2, row):
                cell_value = sheet.cell(row=row_num, column=1).value
                if cell_value is not None and cell_value.lower() == key.lower():
                    found=True
                    sheet = self.add_row(sheet,value,row_num,col,found)
                    break  # Exit the loop after finding a match
            if not found:
                sheet.cell(row=row, column=1, value=key.lower())
                sheet = self.add_row(sheet,value,row,col,found)



            #------
            self.update_col_width(sheet,len(str(value[0])) , col)
            self.update_col_width(sheet,len(self.build)+1 , col)
            self.update_col_width(sheet,len(self.build)+1 , col-1)
            self.update_col_width(sheet,len(str(value[0])) , col-1)
            self.update_col_width(sheet,58 , 1)
            #-----

    def do_modify(self):
        with open(self.current_build_data_path, 'r') as file:
                current_data = json.load(file)
                print("Fetched current build node-wise data")
                
        if os.path.exists(self.prev_build_data_path):
            with open(self.prev_build_data_path, 'r') as file:
                print("Fetched previous build node-wise data")
                prev_data = json.load(file)
            try:prev_data['memory'] 
            except:prev_data["memory"]={}
            try:prev_data['cpu']
            except:prev_data["cpu"]={}
            try:prev_data['overall_memory']
            except:prev_data["overall_memory"]={}
            try:prev_data['overall_cpu']
            except:prev_data["overall_cpu"]={}
            try:prev_data['container_wise_memory']
            except:prev_data["container_wise_memory"]={}
            try:prev_data['container_wise_cpu']
            except:prev_data["container_wise_cpu"]={}
            try:prev_data["complete_usage"]
            except:prev_data["complete_usage"]={}
            try:prev_data["node-apps"]
            except:prev_data["node-apps"]={"memory":{},"cpu":{}}
            
        else:
            print("Previous build node-wise data doesnt exist, created a dummy dict")
            prev_data={
                "memory":{},
                "cpu":{},
                "overall_memory":{},
                "overall_cpu":{},
                "container_wise_memory":{},
                "container_wise_cpu":{},
                "complete_usage":{},
                "node-apps":{"memory":{},"cpu":{}}
            }
        
        current_memory,current_cpu,current_overall_mem, current_overall_cpu = current_data['memory'] , current_data['cpu'],current_data['overall_memory'],current_data['overall_cpu']
        prev_memory,prev_cpu,prev_overall_mem, prev_overall_cpu = prev_data['memory'] , prev_data['cpu'],prev_data['overall_memory'],prev_data['overall_cpu']

        memory_data_dict=self.create_dict_cpu_mem(current_memory,prev_memory,"GB","memory")
        cpu_data_dict=self.create_dict_cpu_mem(current_cpu,prev_cpu,"cores","cpu")
        overall_memory_data_dict=self.create_dict_overall(current_overall_mem,prev_overall_mem,"GB","memory")
        overall_cpu_data_dict=self.create_dict_overall(current_overall_cpu,prev_overall_cpu,"cores","cpu")

        kafka_topics_list = current_data["kafka_topics"]
        kafka_topics_dict=self.create_kafka_dict(kafka_topics_list)


        container_wise_current_memory,container_wise_current_cpu,curr_complete_usage = current_data['container_wise_memory'] , current_data['container_wise_cpu'],current_data["complete_usage"]
        container_wise_prev_memory,container_wise_prev_cpu,prev_complete_usage = prev_data['container_wise_memory'] , prev_data['container_wise_cpu'],prev_data["complete_usage"]

        container_wise_memory_data_dict=self.create_dict_container_wise(container_wise_current_memory,container_wise_prev_memory,"GB","memory")
        container_wise_cpu_data_dict=self.create_dict_container_wise(container_wise_current_cpu,container_wise_prev_cpu,"cores","cpu")
        complete_usage_dict = self.create_complete_dict(curr_complete_usage,prev_complete_usage)

        # curr_node_apps_mem,curr_node_apps_cpu  = current_data["node-apps"]["memory"],current_data["node-apps"]["cpu"]
        # prev_node_apps_mem,prev_node_apps_cpu  = prev_data["node-apps"]["memory"],prev_data["node-apps"]["cpu"]

        # node_apps_memory_data_dict=self.create_dict_node_apps(curr_node_apps_mem,prev_node_apps_mem,"GB","memory")
        # node_apps_cpu_data_dict=self.create_dict_node_apps(curr_node_apps_cpu,prev_node_apps_cpu,"cores","cpu")

        # node_apps_dict = {**node_apps_memory_data_dict, **node_apps_cpu_data_dict}

        sheets = {
                 "memory trend":memory_data_dict,
                  "cpu trend":cpu_data_dict,
                  "container-wise memory trend":container_wise_memory_data_dict,
                  "container-wise cpu trend":container_wise_cpu_data_dict,
                  "overall mem":overall_memory_data_dict,
                  "overall cpu":overall_cpu_data_dict,
                  "complete usage":complete_usage_dict,
                  "kafka topics" : kafka_topics_dict,
                #   "node-apps" : node_apps_dict
                  }

        if not os.path.exists(self.previous_excel_file_path):
            workbook = openpyxl.Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            for sheet in sheets:
                print(f"Added sheet {sheet}")
                sheet1 = workbook.create_sheet(sheet)
            print("created new excel sheet and added new sheets")
        else:
            
            workbook = openpyxl.load_workbook(self.previous_excel_file_path)
            missing_sheets = [sheet_name for sheet_name in sheets if sheet_name not in workbook.sheetnames]
            for sheet_name in missing_sheets:
                print(f"Added sheet {sheet_name}")
                workbook.create_sheet(sheet_name)
            print("Loaded existing excel and added new sheets")
        for sheet in sheets:
            print(f"Processing sheet {sheet} ...")
            self.add_columns(sheet,sheets[sheet],workbook)
        
        workbook.save(self.current_excel_file_path)
        print("Updated Excel sheet succcessfully")
